import os
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font

# Импортируем модули ReportLab для экспорта в PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


# Регистрируем шрифт Calibri (убедитесь, что путь к Calibri.ttf указан правильно)
pdfmetrics.registerFont(TTFont('Calibri', 'Calibri.ttf'))


def process_exception_column(wb, filename, exceptions_list, sheet_name="Balances", header_row=2):
    """
    Если имя файла (filename) содержит одно из слов из exceptions_list (без учёта регистра),
    то для листа sheet_name (по умолчанию "Balances") выполняется следующее:
      1. Находится столбец с заголовком "IM" в строке заголовков (header_row).
      2. Если сразу после него (то есть в столбце с индексом IM+1) уже не стоит заголовок "Trading credits",
         то вставляется новый столбец в эту позицию.
      3. В новой колонке в заголовке устанавливается "Trading credits".
      4. Для всех строк с данными (начиная со строки header_row+1) в этой колонке устанавливается значение 1,000,000.00.
      5. Затем для каждой строки пересчитывается значение в колонке "Equity" как:
             Equity = Margin Balance - Trading credits
         (учтите, что после вставки столбец "Equity" смещается на одну позицию вправо, если он находился правее "IM").
    """
    # Проверяем наличие исключения в имени файла (без учёта регистра)
    if not any(exc.lower() in filename.lower() for exc in exceptions_list):
        return  # Если имя файла не содержит ни одного из исключений, выходим

    if sheet_name not in wb.sheetnames:
        print(f"Лист '{sheet_name}' не найден в книге.")
        return

    ws = wb[sheet_name]

    # Считываем заголовочную строку (header_row)
    header_cells = list(ws[header_row])

    im_index = None
    margin_index = None
    equity_index = None

    # Определяем номера столбцов по заголовкам (сравниваем без учета регистра и пробелов)
    for cell in header_cells:
        if cell.value is not None:
            val = str(cell.value).strip().lower()
            if val == "im":
                im_index = cell.column  # 1-индексация
            elif val == "margin balance":
                margin_index = cell.column
            elif val == "equity":
                equity_index = cell.column

    if im_index is None:
        print("Столбец 'IM' не найден в заголовке листа.")
        return
    if margin_index is None:
        print("Столбец 'Margin Balance' не найден в заголовке листа.")
        return
    if equity_index is None:
        print("Столбец 'Equity' не найден в заголовке листа.")
        return

    # Проверяем, не существует ли уже столбец "Trading credits" сразу после "IM".
    # Если заголовок в ячейке с индексом (im_index + 1) равен "Trading credits", то ничего не делаем.
    next_cell = ws.cell(row=header_row, column=im_index + 1).value
    if next_cell is not None and str(next_cell).strip().lower() == "trading credits":
        print("Столбец 'Trading credits' уже существует. Обновляем его данные...")
        trading_credits_index = im_index + 1
    else:
        # Вставляем новый столбец после "IM" (то есть в позиции im_index + 1)
        ws.insert_cols(im_index + 1)
        # Записываем заголовок "Trading credits" в строке заголовков
        ws.cell(row=header_row, column=im_index + 1, value="Trading credits")
        trading_credits_index = im_index + 1
        # Обновляем индексы столбцов для тех, что находятся правее новой колонки.
        # Если Equity находился правее "IM", то его новый индекс увеличивается на 1.
        if equity_index > im_index:
            equity_index += 1

    # # Записываем значение Trading credits (1,000,000.00) для каждой строки с данными (начиная со строки header_row+1)
    for r in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row=r, column=trading_credits_index)
        cell.value = 0.00
        cell.number_format = "#,##0.00"

    # Пересчитываем Equity для каждой строки:
    # Equity = Margin Balance - Trading credits
    for r in range(header_row + 1, ws.max_row + 1):
        try:
            margin_val = float(ws.cell(row=r, column=margin_index).value)
        except (TypeError, ValueError):
            margin_val = 0.0
        try:
            equity_val = float(ws.cell(row=r, column=equity_index).value)
        except (TypeError, ValueError):
            equity_val = 0.0

        new_trading_val = margin_val - equity_val if margin_val > 0 else 0
        tc_cell = ws.cell(row=r, column=trading_credits_index)
        tc_cell.value = new_trading_val
        tc_cell.number_format = "#,##0.00"

    print(f"Для файла '{filename}' в листе '{sheet_name}' вставлен столбец 'Trading credits' после 'IM', и 'Equity' пересчитана.")


def compute_aggregated_row(ws, header_rows_count, data_start_row, numeric_headers,
                           indicator_header="User ID", aggregate_all=False):
    """
    Для рабочего листа ws:
      - header_rows_count – количество строк, занятых заголовками (например, 2 для Balances, 3 для Trading Summary).
      - data_start_row – номер строки, с которой начинаются данные (обычно header_rows_count + 1).
      - numeric_headers – список имен столбцов для агрегирования (используется, если aggregate_all==False).
      - indicator_header – имя столбца, по которому определяется наличие агрегированной строки.
        Если столбец объединён (например, в Trading Summary), функция ищет его во всех строках заголовочного блока.
      - aggregate_all – если True, агрегируются данные по всем столбцам (кроме индикаторного), иначе только по numeric_headers.

    Алгоритм:
      1. Поиск индикаторного столбца (обычно "User ID") в строках от 1 до header_rows_count.
      2. Если в последней строке (в этом столбце) уже написано "Aggregated" (без учёта регистра) – агрегированная строка уже есть.
         Иначе добавляется новая строка, и в индикаторном столбце записывается "Aggregated".
      3. Если aggregate_all==True, то для каждого столбца (от 1 до ws.max_column, пропуская индикаторный) вычисляется сумма значений
         (от data_start_row до агрегированной строки-1) и записывается в соответствующую ячейку агрегированной строки.
         Если aggregate_all==False, то для каждого имени из numeric_headers ищется его столбец (в заголовочном блоке) и суммируются данные.
    """
    # 1. Поиск индикаторного столбца в заголовочных строках
    indicator_col = None
    for r in range(1, header_rows_count + 1):
        for cell in ws[r]:
            if cell.value is not None and str(cell.value).strip().lower() == indicator_header.lower():
                indicator_col = cell.column
                break
        if indicator_col is not None:
            break
    if indicator_col is None:
        indicator_col = 1  # если не найдено, используем первый столбец

    # 2. Определяем агрегированную строку: если в последней строке в индикаторном столбце уже написано "Aggregated" – используем её;
    # иначе добавляем новую строку.
    last_row = ws.max_row
    cell_val = ws.cell(row=last_row, column=indicator_col).value
    if cell_val is not None and str(cell_val).strip().lower() == "aggregated":
        aggregated_row = last_row
    else:
        aggregated_row = last_row + 1
        ws.cell(row=aggregated_row, column=indicator_col).value = "Aggregated"

    # 3. Агрегирование данных:
    if aggregate_all:
        # Суммируем данные по всем столбцам (от 1 до ws.max_column), пропуская индикаторный столбец.
        for col in range(1, ws.max_column + 1):
            if col == indicator_col:
                continue
            total = 0
            for r in range(data_start_row, aggregated_row):
                val = ws.cell(row=r, column=col).value
                try:
                    total += float(val)
                except (TypeError, ValueError):
                    pass
            agg_cell = ws.cell(row=aggregated_row, column=col)
            agg_cell.value = total
            agg_cell.number_format = "#,##0.00"
    else:
        # Для каждого столбца из numeric_headers ищем его в заголовочном блоке (во всех строках от 1 до header_rows_count)
        for header in numeric_headers:
            col_index = None
            for r in range(1, header_rows_count + 1):
                for cell in ws[r]:
                    if cell.value is not None and str(cell.value).strip() == header:
                        col_index = cell.column
                        break
                if col_index is not None:
                    break
            if col_index is None:
                continue
            total = 0
            for r in range(data_start_row, aggregated_row):
                val = ws.cell(row=r, column=col_index).value
                try:
                    total += float(val)
                except (TypeError, ValueError):
                    pass
            agg_cell = ws.cell(row=aggregated_row, column=col_index)
            agg_cell.value = total
            agg_cell.number_format = "#,##0.00"


# =============================================================
# Функция для загрузки данных из raw-файла и группировки по категориям
# =============================================================
def load_raw_data(raw_file, categories):
    """
    Читает raw-файл и распределяет строки по категориям.
    Строка, в которой первая ячейка равна имени категории, служит разделителем и не сохраняется.
    """
    raw_wb = load_workbook(raw_file, data_only=True)
    raw_ws = raw_wb.active

    data_dict = {cat: [] for cat in categories.keys()}
    current_category = None

    for row in raw_ws.iter_rows(values_only=True):
        first_cell = row[0]  # предполагаем, что в первой ячейке может быть название категории
        if first_cell in categories:
            current_category = first_cell
            continue  # не сохраняем строку с названием категории
        if current_category is not None:
            if all(cell is None for cell in row):
                continue  # пропускаем пустые строки
            data_dict[current_category].append(list(row))
    return data_dict

# =============================================================
# Функция для заполнения шаблона данными и применения стилей/форматирования
# =============================================================
def fill_template(template_file, data_dict, categories, numeric_fields, header_font):
    """
    Загружает шаблонный файл, для каждого листа:
      - записывает данные (начиная со строки после заголовков),
      - применяет стиль к заголовкам,
      - форматирует числовые столбцы (приводит значения к float и задаёт формат "#,##0.00").
    Возвращает изменённую книгу.
    """
    wb = load_workbook(template_file)

    for cat, header_rows in categories.items():
        if cat not in wb.sheetnames:
            print(f"Лист '{cat}' не найден в шаблоне. Пропускаем.")
            continue

        ws = wb[cat]
        start_row = header_rows + 1  # данные записываются после заголовков

        # Запись данных в лист
        for i, data_row in enumerate(data_dict[cat], start=start_row):
            for j, value in enumerate(data_row, start=1):
                ws.cell(row=i, column=j, value=value)

        # Применяем стиль к ячейкам заголовков (строки 1...header_rows)
        for row in range(1, header_rows + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    cell.font = header_font

        # Определяем номера столбцов для числового форматирования по заголовкам
        numeric_cols = []
        header_row_idx = header_rows  # предполагается, что заголовки во последней строке блока
        for cell in ws[header_row_idx]:
            if cell.value in numeric_fields.get(cat, []):
                numeric_cols.append(cell.column)

        # Приводим значения к числовому типу и применяем формат "#,##0.00"
        for col in numeric_cols:
            for row in range(start_row, start_row + len(data_dict[cat])):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    try:
                        numeric_val = float(cell.value)
                        cell.value = numeric_val
                        cell.number_format = "#,##0.00"
                    except (ValueError, TypeError):
                        pass

        print(f"Лист '{cat}' заполнен и стилизован (данные начинаются с {start_row}-й строки).")

    return wb

# =============================================================
# Функция для удаления лишних строк с дублирующими заголовками
# =============================================================
def delete_duplicate_headers(workbook, deletion_rules):
    """
    Удаляет заданные строки с каждого листа, согласно deletion_rules.
    deletion_rules — словарь вида:
         { "Имя листа": (начальная_строка, количество_удаляемых_строк) }
    """
    for sheet_name, (start_row, num_rows) in deletion_rules.items():
        if sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            ws.delete_rows(start_row, num_rows)
            print(f"Удалены строки {start_row} - {start_row + num_rows - 1} в листе '{sheet_name}'.")
        else:
            print(f"Лист '{sheet_name}' не найден в книге для удаления строк.")

# =============================================================
# Функция для экспорта данных из книги в PDF-файл
# =============================================================
def export_to_pdf(workbook, sheet_order, output_pdf_file, header_rows_pdf, cover_company=None):
    """
    Экспортирует содержимое листов workbook в PDF-файл.
    Для каждого листа:
      - Каждая ячейка оборачивается в Paragraph для переноса длинного текста.
      - Для первых N строк (N задаётся в header_rows_pdf для данного листа) используется header_style
        с выравниванием по центру (горизонтально и вертикально), для остальных – body_style.
      - Обрабатываются объединённые ячейки (merged cells) с помощью команды SPAN.
      - Таблица растягивается равномерно на всю ширину страницы, а число столбцов определяется по первой (заголовочной) строке,
        при этом если пустые ячейки являются частью объединённых диапазонов, они учитываются.
    Параметр header_rows_pdf – словарь вида: { "Лист": число_заголовочных_строк }.
    """

    # Определяем стили для заголовков и для остальных ячеек
    header_style = ParagraphStyle(
        name="header_style",
        fontName="Calibri",
        fontSize=6,
        leading=10,
        alignment=TA_CENTER
    )
    body_style = ParagraphStyle(
        name="body_style",
        fontName="Calibri",
        fontSize=4,
        leading=6,
        alignment=TA_CENTER  # левое выравнивание для остальных ячеек
    )

    # Настраиваем документ PDF
    doc = SimpleDocTemplate(
        output_pdf_file,
        pagesize=A4,
        leftMargin=20,
        rightMargin=20,
        topMargin=20,
        bottomMargin=20
    )
    elements = []
    styles = getSampleStyleSheet()
    # title_style = styles["Heading1"]
    # title_style.alignment = TA_CENTER  # Заголовок листа по центру

    title_style = ParagraphStyle(
        name='MyTitleStyle',
        fontName='Calibri',  # Используем Calibri
        fontSize=10,  # Размер шрифта 14
        leading=6,  # Межстрочный интервал (можно настроить)
        alignment=0  # Выравнивание по центру
    )

    # Если задан cover_company, то добавляем первую обложную страницу
    if cover_company:
        # Определяем стили для обложной страницы
        cover_style1 = ParagraphStyle(
            name="cover_style1",
            fontName="Calibri",
            fontSize=24,
            leading=28,
            alignment=TA_CENTER
        )
        cover_style2 = ParagraphStyle(
            name="cover_style2",
            fontName="Calibri",
            fontSize=18,
            leading=22,
            alignment=TA_CENTER
        )
        cover_style3 = ParagraphStyle(
            name="cover_style3",
            fontName="Calibri",
            fontSize=12,
            leading=16,
            alignment=TA_CENTER
        )
        # Формируем строки обложной страницы
        cover_text1 = Paragraph("MONTHLY STATEMENT", cover_style1)
        cover_text2 = Paragraph("COINCALL X " + cover_company.upper(), cover_style2)
        utc_now = datetime.datetime.utcnow().strftime("%d %b %Y %H:%M UTC")
        cover_text3 = Paragraph("REPORT TIME: " + utc_now.upper(), cover_style3)
        # Добавляем элементы: между строками можно задать отступы
        elements.append(Spacer(1, 300))
        elements.append(cover_text1)
        elements.append(Spacer(1, 20))
        elements.append(cover_text2)
        elements.append(Spacer(1, 20))
        elements.append(cover_text3)
        elements.append(Spacer(1, 100))
        elements.append(PageBreak())  # Разрыв страницы: таблицы начнутся со второй страницы

    # Обрабатываем листы в порядке, заданном в sheet_order
    for sheet_name in sheet_order:
        if sheet_name not in workbook.sheetnames:
            print(f"Лист '{sheet_name}' не найден для экспорта в PDF.")
            continue

        ws = workbook[sheet_name]
        data = []  # Будущий список строк для таблицы PDF
        excel_to_pdf_index = {}  # Сопоставление: номер строки Excel -> индекс строки в data

        # Количество заголовочных строк для данного листа (если не указано – 1)
        header_count = header_rows_pdf.get(sheet_name, 1)

        # Формируем data: включаем все строки, где хотя бы одна ячейка не пуста
        for row in ws.rows:
            if any(cell.value is not None for cell in row):
                pdf_index = len(data)
                excel_to_pdf_index[row[0].row] = pdf_index
                row_data = []
                for cell in row:
                    if cell.value is None:
                        text = ""
                    else:
                        # Если ячейка содержит число с форматом "#,##0.00", форматируем его
                        if isinstance(cell.value, (int, float)) and cell.number_format == "#,##0.00":
                            text = f"{cell.value:,.2f}"
                        else:
                            text = str(cell.value)
                    # Для первых header_count строк используем стиль заголовка, для остальных – базовый стиль
                    style_to_use = header_style if pdf_index < header_count else body_style
                    para = Paragraph(text, style_to_use)
                    row_data.append(para)
                data.append(row_data)

        if not data:
            continue

        # Определяем Excel-номер строки для заголовка (первая строка, попавшая в PDF)
        header_excel_row = None
        for ex_row, pdf_index in excel_to_pdf_index.items():
            if pdf_index == 0:
                header_excel_row = ex_row
                break
        if header_excel_row is None:
            header_excel_row = 1

        # --- Вычисление "эффективного" числа столбцов по первой (заголовочной) строке ---
        header_row = data[0]
        effective_max_cols = 0
        # Перебираем столбцы с конца к началу
        for col in range(len(header_row), 0, -1):
            text = header_row[col - 1].getPlainText().strip()
            if text != "":
                effective_max_cols = col
                break
            else:
                # Если ячейка пустая, проверяем, входит ли она в объединённый диапазон с ненулевым значением
                for merged_range in ws.merged_cells.ranges:
                    if (merged_range.min_row <= header_excel_row <= merged_range.max_row and
                            merged_range.min_col <= col <= merged_range.max_col):
                        # Получаем значение из верхней левой ячейки объединённого диапазона
                        cell_val = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
                        if cell_val is not None and str(cell_val).strip() != "":
                            effective_max_cols = col
                            break
                if effective_max_cols:
                    break
        if effective_max_cols == 0:
            effective_max_cols = len(header_row)
        max_cols = effective_max_cols

        # Дополняем каждую строку до max_cols, если она короче; если длиннее – обрезаем лишнее
        for idx, r in enumerate(data):
            if len(r) < max_cols:
                for _ in range(max_cols - len(r)):
                    r.append(Paragraph("", body_style))
            elif len(r) > max_cols:
                data[idx] = r[:max_cols]

        # Равномерно распределяем ширину столбцов по всей доступной ширине страницы
        col_width = doc.width / max_cols if max_cols else doc.width
        col_widths = [col_width] * max_cols

        # Создаём объект таблицы
        table = Table(data, colWidths=col_widths) # , rowHeights=10

        # Базовые команды стиля таблицы: сетка, выравнивание, фон и шрифт для заголовка (первые header_count строк)
        table_style_commands = [
            ('GRID', (0, 0), (-1, -1), 0.1, colors.gray),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, 0), (-1, header_count - 1), colors.lightgrey),
            ('FONTNAME', (0, 0), (-1, header_count - 1), 'Calibri'), # Helvetica-Bold
            ('FONTSIZE', (0, 0), (-1, -1), 5),
        ]

        # Обработка объединённых ячеек (merged cells) из Excel
        for merged_range in ws.merged_cells.ranges:
            min_row = merged_range.min_row
            max_row = merged_range.max_row
            min_col = merged_range.min_col
            max_col = merged_range.max_col
            if min_row in excel_to_pdf_index and max_row in excel_to_pdf_index:
                pdf_min_row = excel_to_pdf_index[min_row]
                pdf_max_row = excel_to_pdf_index[max_row]
                pdf_min_col = min_col - 1  # перевод в 0-индексацию
                pdf_max_col = max_col - 1
                # Если объединение выходит за пределы max_cols, обрезаем его
                if pdf_min_col >= max_cols:
                    continue
                pdf_max_col = min(pdf_max_col, max_cols - 1)
                if pdf_max_col < pdf_min_col:
                    continue
                table_style_commands.append(('SPAN', (pdf_min_col, pdf_min_row), (pdf_max_col, pdf_max_row)))

        # Обработка фонового цвета ячеек из Excel
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            if row[0].row not in excel_to_pdf_index:
                continue
            pdf_row = excel_to_pdf_index[row[0].row]
            for j, cell in enumerate(row, start=1):
                if j > max_cols:
                    continue
                fill = cell.fill
                if fill is not None and fill.fill_type == "solid":
                    fgColor = fill.fgColor
                    if fgColor is not None and fgColor.rgb is not None:
                        rgb = fgColor.rgb
                        if len(rgb) == 8:
                            rgb = rgb[-6:]  # удаляем альфа-канал, если есть
                        try:
                            cell_color = colors.HexColor("#" + rgb)
                            table_style_commands.append(('BACKGROUND', (j - 1, pdf_row), (j - 1, pdf_row), cell_color))
                        except Exception:
                            pass

        table.setStyle(TableStyle(table_style_commands))

        # Добавляем в PDF заголовок листа, отступ и затем таблицу
        elements.append(Paragraph(sheet_name.upper(), title_style))
        elements.append(Spacer(1, 12))
        elements.append(table)
        elements.append(Spacer(1, 24))

    doc.build(elements)
    print(f"PDF-файл успешно сохранён: {output_pdf_file}")

# =============================================================================
# Главная функция для обработки всех файлов в директории raw_data
# =============================================================================
def process_all_raw_files():
    # Пути к папкам
    base_folder = r"C:\Users\_\Desktop\CC_report"
    raw_data_dir = os.path.join(base_folder, "raw_data")
    template_file = os.path.join(base_folder, "template", "temp1.xlsx")
    result_folder = os.path.join(base_folder, "result")
    os.makedirs(result_folder, exist_ok=True)

    # Настройки: категории, числовые поля, правила удаления, количество заголовочных строк для PDF
    categories = {
        "Balances": 2,
        "Trading Summary": 3,
        "Fees Summary": 3,
        "Positions": 2,
    }
    numeric_fields = {
        "Balances": [
            "Margin Balance", "MM", "IM", "Equity",
            "Available Balance", "Monthly Deposit",
            "Monthly Withdrawals", "Monthly Net Deposit"
        ],
        "Trading Summary": [
            "Total Volume", "Taker Volume", "Maker Volume"
        ],
        "Fees Summary": [
            "Total Fees", "Taker Fees", "Maker Fees"
        ],
        "Positions": [
            "Size", "Value", "Index", "Mark"
        ],
    }
    deletion_rules = {
        "Balances": (3, 2),
        "Trading Summary": (4, 3),
        "Fees Summary": (4, 3),
        "Positions": (3, 2),
    }
    header_rows_pdf = {
        "Balances": 2,
        "Trading Summary": 3,
        "Fees Summary": 3,
        "Positions": 2,
    }
    sheet_order = ["Balances", "Trading Summary", "Fees Summary", "Positions"]
    header_font = Font(bold=True, size=11)

    # Обрабатываем каждый файл в директории raw_data
    for filename in os.listdir(raw_data_dir):
        if not filename.lower().endswith('.xlsx'):
            continue
        raw_file = os.path.join(raw_data_dir, filename)
        print("Обработка файла:", raw_file)
        data_dict = load_raw_data(raw_file, categories)
        wb = fill_template(template_file, data_dict, categories, numeric_fields, header_font)
        delete_duplicate_headers(wb, deletion_rules)

        # Массив исключений (на данный момент только "Antalpha")
        exceptions_list = ["Antalpha"]

        # Если имя файла содержит одно из исключённых слов, обрабатываем лист Balances
        process_exception_column(wb, filename, exceptions_list, sheet_name="Balances", header_row=categories["Balances"])

        # Вычисление агрегированных строк для листов
        if "Balances" in wb.sheetnames:
            ws = wb["Balances"]
            header_row_num = categories["Balances"]
            data_start_row = header_row_num + 1
            compute_aggregated_row(ws, header_row_num, data_start_row, numeric_fields["Balances"], indicator_header="User ID", aggregate_all=False)
        if "Trading Summary" in wb.sheetnames:
            ws = wb["Trading Summary"]
            header_row_num = categories["Trading Summary"]
            data_start_row = header_row_num + 1
            compute_aggregated_row(ws, header_row_num, data_start_row, numeric_fields["Trading Summary"], indicator_header="User ID", aggregate_all=True)
        if "Fees Summary" in wb.sheetnames:
            ws = wb["Fees Summary"]
            header_row_num = categories["Fees Summary"]
            data_start_row = header_row_num + 1
            compute_aggregated_row(ws, header_row_num, data_start_row, numeric_fields["Fees Summary"], indicator_header="User ID", aggregate_all=True)

        # Формируем новое имя файла по схеме:
        # Из исходного имени (например, "mm-monthly-report-Orbit.xlsx") берём последнюю часть "Orbit"
        base_name = os.path.splitext(filename)[0]
        parts = base_name.split("-")
        report_name = parts[-1].strip().capitalize() if parts else base_name.capitalize()
        # Вычисляем сокращённое имя предыдущего месяца (например, если сегодня февраль 2025, то предыдущий месяц - Jan25)
        today = datetime.date.today()
        first_day_this_month = today.replace(day=1)
        last_day_prev_month = first_day_this_month - datetime.timedelta(days=1)
        suffix = last_day_prev_month.strftime("%b") + last_day_prev_month.strftime("%y")
        new_base_name = f"{report_name} Coincall Monthly Report {suffix}"

        # Сохраняем Excel-файл
        new_xlsx_file = os.path.join(result_folder, new_base_name + ".xlsx")
        wb.save(new_xlsx_file)
        print("Сохранён Excel-файл:", new_xlsx_file)
        # Экспорт в PDF
        new_pdf_file = os.path.join(result_folder, new_base_name + ".pdf")
        export_to_pdf(wb, sheet_order, new_pdf_file, header_rows_pdf, cover_company=report_name)
        print("Сохранён PDF-файл:", new_pdf_file)

# =============================================================
# Точка входа в программу
# =============================================================
if __name__ == "__main__":
    process_all_raw_files()
